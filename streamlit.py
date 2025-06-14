import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import altair as alt
from datetime import datetime
from io import BytesIO

# Função para processar os dados com base no tipo de DataLogger
def processar_dados(tipo_datalogger, arquivo):
    try:
        # Lendo o arquivo enviado
        extensao = os.path.splitext(arquivo.name)[-1]
        if extensao == ".csv":
            df = pd.read_csv(arquivo, sep=";", skiprows=5)
        elif extensao in [".xls", ".xlsx"]:
            df = pd.read_excel(arquivo)

        # Processamento baseado no tipo de DataLogger
        if tipo_datalogger == "Lamon":
            # if 'Hora' in df.columns:
            #     df['Hora'], df['Minutos'], df['Segundos'] = zip(
            #         *df['Hora'].str.split(':').apply(lambda x: (x[0].zfill(2), x[1].zfill(2), x[2].zfill(2)))
            #     )
            if 'Data' in df.columns and 'Hora' in df.columns:
                df['DataHora'] = pd.to_datetime(df['Data'] + ' ' + df['Hora'], dayfirst=True, errors='coerce')
                df.drop(columns=['Data', 'Hora'], inplace=True) 

            if 'Pressão(mca)' in df.columns:
                df['Pressão(mca)'] = df['Pressão(mca)'].astype(float).round(2)
                df['Pressão_mca'] = df['Pressão(mca)'].apply(lambda x: f"{x:.2f}")
                df.drop(columns=['Pressão(mca)'], inplace=True)

        elif tipo_datalogger == "Sanesoluti":
            # Converter a coluna 'Data' para datetime se ainda não estiver no formato
            df['Data'] = pd.to_datetime(df['Data'])
            # Converter a coluna 'Hora' para timedelta (para adicionar à data)
            df['Hora'] = pd.to_timedelta(df['Hora'])
            df['DataHora'] = df['Data'] + df['Hora']
            # if 'Hora' in df.columns:
            #     # Dividir a coluna 'Hora' em Hora, Minutos e Segundos
            #     df['Hora'], df['Minutos'], df['Segundos'] = zip(
            #         *df['Hora'].str.split(':').apply(lambda x: (x[0].zfill(2), x[1].zfill(2), x[2].zfill(2)))
            #     )

            if 'Pressão' in df.columns and 'Volume Total' in df.columns:
                # Ajustar os valores de Pressão e Volume Total
                df['Pressão'] = df['Pressão'].astype(float).round(2)
                df['Volume Total'] = df['Volume Total'].astype(float).round(2)

        elif tipo_datalogger == "Sanesoluti V.2":
    # Converter a coluna 'Data' para datetime se ainda não estiver no formato
            df['Data'] = pd.to_datetime(df['Data'], dayfirst=True)
            
            # Converter a coluna 'Hora' de fração do dia para timedelta
            df['Hora'] = pd.to_numeric(df['Hora'], errors='coerce')  # Garante que é float
            df['Hora'] = pd.to_timedelta(df['Hora'] * 24, unit='h')  # Multiplica por 24 para transformar em horas
            
            df['DataHora'] = df['Data'] + df['Hora']

            if 'Pressão' in df.columns and 'Volume G' in df.columns:
                # Ajustar os valores de Pressão e Volume Total
                df['Pressão'] = df['Pressão'].astype(float).round(2)
                df['Volume G'] = df['Volume G'].astype(float).round(2)

        return df


    except Exception as e:
        st.error(f"Erro ao processar os dados: {e}")
        return None

# Função para converter DataFrame em Excel com tabela formatada
def converter_para_excel_com_tabela(df):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active

    # Adicionar cabeçalhos na planilha
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Adicionar dados do DataFrame na planilha
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Criar tabela formatada
    ref = f"A1:{chr(64 + len(df.columns))}{len(df) + 1}"  # Define o intervalo da tabela (ex: A1:C4)
    tabela = Table(displayName="Tabela1", ref=ref)
    estilo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)

    # Salvar o arquivo na memória
    wb.save(output)
    return output.getvalue()

# Interface do Streamlit
st.title("Processador de Arquivos DataLogger")

# Upload do arquivo
arquivo = st.file_uploader("Faça upload do arquivo (.csv, .xls, .xlsx)", type=["csv", "xls", "xlsx"])

# Seleção do tipo de DataLogger
tipos_datalogger = ["Lamon", "Sanesoluti", "Sanesoluti V.2" ,"Vectora"]

tipo_datalogger = st.selectbox("Selecione o tipo de DataLogger", tipos_datalogger)
if arquivo is not None and tipo_datalogger:
    resultado_df = processar_dados(tipo_datalogger, arquivo)

    d = st.date_input(
        "Select your vacation for next year",
        (min(resultado_df["DataHora"]),max(resultado_df["DataHora"])),
        min(resultado_df["DataHora"]),
        max(resultado_df["DataHora"]),
        format="DD/MM/YYYY",
    )
    # Verificar se o widget retornou um intervalo válido
    start_date, end_date = None, None
    # Filtrar o DataFrame com base no intervalo selecionado
    if isinstance(d, tuple) and len(d) == 2:
        start_date = d[0]
        end_date = d[1]
    if start_date and end_date :
        resultado_df = resultado_df[
            (resultado_df['DataHora'].dt.date >= start_date) &
            (resultado_df['DataHora'].dt.date <= end_date)
        ]

if st.button("Processar"):
       # Define a data padrão (por exemplo, a data atual)

    if resultado_df is not None:
            st.dataframe(resultado_df)

            # Exibir os dados processados
            st.success("Dados processados com sucesso!")
            if tipo_datalogger == "Lamon":
                # Verificar se as colunas necessárias estão presentes no DataFrame para gerar o gráfico
                if 'DataHora' in resultado_df.columns and 'Pressão_mca' in resultado_df.columns:
                    # Converter a coluna 'Data' para datetime se necessário (ajuste conforme o formato dos seus dados)
                    resultado_df['DataHora'] = pd.to_datetime(resultado_df['DataHora'], errors='coerce')
                    resultado_df['Pressão_mca'] = pd.to_numeric(resultado_df['Pressão_mca'], errors='coerce')
                    resultado_df = resultado_df.sort_values(by='DataHora')
                    
                    # Criar gráfico Altair
                    chart = alt.Chart(resultado_df).mark_line().encode(
                        x=alt.X('DataHora:T', title='Data e Hora'),
                        y=alt.Y('Pressão_mca:Q', title='Pressão (mca)'),
                        tooltip=['DataHora', 'Pressão_mca']
                    ).properties(
                        width=500,
                        height=600,
                        # title="Gráfico de Pressão ao Longo do Tempo"
                    )
                    # Renderizar gráfico no Streamlit
                    st.altair_chart(chart, use_container_width=True)


                else:
                    st.warning("As colunas 'Data' e/ou 'Pressão' não estão presentes no arquivo enviado.")
            
            elif tipo_datalogger == "Sanesoluti":
                # Verificar se as colunas necessárias estão presentes no DataFrame para gerar o gráfico
                if 'DataHora' in resultado_df.columns and 'Pressão' in resultado_df.columns and 'Volume Total' in resultado_df.columns:
                    # Converter a coluna 'Data' para datetime se necessário (ajuste conforme o formato dos seus dados)
                    resultado_df['DataHora'] = pd.to_datetime(resultado_df['DataHora'], errors='coerce')
                    resultado_df['Pressão'] = pd.to_numeric(resultado_df['Pressão'], errors='coerce')
                    resultado_df['Volume Total'] = pd.to_numeric(resultado_df['Volume Total'], errors='coerce')
                    resultado_df = resultado_df.sort_values(by='DataHora')
                    # Criar gráfico Altair# Reformatar o DataFrame para múltiplas linhas (long format)
                      # Criar gráfico Altair
                    chart = alt.Chart(resultado_df).mark_line().encode(
                        x=alt.X('DataHora:T', title='Data e Hora'),
                        y=alt.Y('Pressão:Q', title='Pressão (mca)'),
                        tooltip=['DataHora', 'Pressão']
                    ).properties(
                        width=500,
                        height=600,
                        # title="Gráfico de Pressão ao Longo do Tempo"
                    )
                    # Renderizar gráfico no Streamlit
                    st.altair_chart(chart, use_container_width=True)
            # elif tipo_datalogger == "Sanesoluti V.2":
            #     # Verificar se as colunas necessárias estão presentes no DataFrame para gerar o gráfico
                
            #     if 'DataHora' in resultado_df.columns and 'Pressão ' in resultado_df.columns and 'Volume G ' in resultado_df.columns:
            #         resultado_df['DataHora'] = pd.to_datetime(resultado_df['DataHora'], dayfirst=True, errors='coerce')
            #         resultado_df['Pressão '] = pd.to_numeric(resultado_df['Pressão '], errors='coerce')
            #         resultado_df['Volume G '] = pd.to_numeric(resultado_df['Volume G '], errors='coerce')
            #         resultado_df = resultado_df.sort_values(by='DataHora')

            #         # Formatando o eixo X para exibir dia e hora
            #         chart = alt.Chart(resultado_df).mark_line().encode(
            #             x=alt.X('DataHora:T', title='Data e Hora', axis=alt.Axis(format="%d/%m %H:%M")),
            #             y=alt.Y('Pressão :Q', title='Pressão (mca)'),
            #             tooltip=['DataHora', 'Pressão ']
            #         ).properties(
            #             width=500,
            #             height=600,
            #         )

            #         st.altair_chart(chart, use_container_width=True)
            elif tipo_datalogger == "Sanesoluti V.2":
                if 'DataHora' in resultado_df.columns and 'Pressão ' in resultado_df.columns and 'Volume G ' in resultado_df.columns:
                    # Pré-processamento dos dados
                    resultado_df['DataHora'] = pd.to_datetime(resultado_df['DataHora'], dayfirst=True, errors='coerce')
                    resultado_df['Pressão '] = pd.to_numeric(resultado_df['Pressão '], errors='coerce').dropna()
                    resultado_df['Volume G '] = pd.to_numeric(resultado_df['Volume G '], errors='coerce').dropna()
                    resultado_df = resultado_df.sort_values(by='DataHora')

                    # Cálculo dos limites do eixo Y
                    min_pressao = resultado_df['Pressão '].min()
                    max_pressao = resultado_df['Pressão '].max()

                    # Criação do gráfico
                    chart = alt.Chart(resultado_df).mark_line(point=True).encode(
                        x=alt.X('DataHora:T',
                            title='Data e Hora',
                            axis=alt.Axis(
                                format="%d/%m %H:%M",
                                labelAngle=-90,  # Rótulos na vertical
                                tickCount=72
                            )
                        ),
                        y=alt.Y('Pressão :Q', 
                            title='Pressão (mca)',
                            scale=alt.Scale(domain=[min_pressao, max_pressao])),
                        tooltip=[alt.Tooltip('DataHora:T', format='%d/%m/%Y %H:%M'), 'Pressão ']
                    ).properties(
                        width=500,
                        height=600
                    )
                    
                    st.altair_chart(chart, use_container_width=True)


                else:
                    st.warning("As colunas 'DataHora', 'Pressão_mca' e/ou 'Volume Total' não estão presentes no arquivo enviado.")
            if tipo_datalogger == "Lamon" and 'Pressão_mca' in resultado_df.columns:
                            min_val = resultado_df['Pressão_mca'].min()
                            max_val = resultado_df['Pressão_mca'].max()
                            mean_val = resultado_df['Pressão_mca'].mean()

                            # Exibir os valores calculados
                            st.write("### Estatísticas da Pressão (mca)")
                            st.write(f"**Valor Mínimo:** {min_val:.2f}")
                            st.write(f"**Valor Médio:** {mean_val:.2f}")
                            st.write(f"**Valor Máximo:** {max_val:.2f}")

            elif tipo_datalogger == "Sanesoluti" and 'Pressão' in resultado_df.columns:
                min_val = resultado_df['Pressão'].min()
                max_val = resultado_df['Pressão'].max()
                mean_val = resultado_df['Pressão'].mean()

                # Exibir os valores calculados
                st.write("### Estatísticas da Pressão")
                st.write(f"**Valor Mínimo:** {min_val:.2f}")
                st.write(f"**Valor Médio:** {mean_val:.2f}")
                st.write(f"**Valor Máximo:** {max_val:.2f}")
            
            elif tipo_datalogger == "Sanesoluti V.2" and 'Pressão ' in resultado_df.columns:
                min_val = resultado_df['Pressão '].min()
                max_val = resultado_df['Pressão '].max()
                mean_val = resultado_df['Pressão '].mean()

                # Exibir os valores calculados
                st.write("### Estatísticas da Pressão")
                st.write(f"**Valor Mínimo:** {min_val:.2f}")
                st.write(f"**Valor Médio:** {mean_val:.2f}")
                st.write(f"**Valor Máximo:** {max_val:.2f}")
    else:
        st.error("Por favor, faça upload de um arquivo e selecione um tipo de DataLogger.")
